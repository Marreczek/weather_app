import os
from openpyxl import Workbook, load_workbook

def save_to_excel(filename, data):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active # wybranie aktywnego arkusza
    else:
        wb = Workbook() # stworzenie nowego excela
        ws = wb.active # wybranie aktywnego arkusza
        ws.append(list(data.keys()))

    ws.append(list(data.values()))
    wb.save(filename)




# def create_excel():
#     wb = Workbook()
#     sheet = wb.active
#     sheet.title = "Temperatury"
#
#     sheet.append(["Pomiar nr", "Temperatura", "Test"])
#     sheet.append([1, 23, "test1"])
#     sheet.append([2, 31, "test2"])
#
#     wb.save("testowy.xlsx")
#
# create_excel()